# dags/healthcare_dataload_dag.py
"""
healthcare_dataload_dag

Pipeline:
1) Download XLSX from Mass.gov (Playwright)
2) Convert XLSX -> cleaned CSV (drop top 8 rows)
3) Upload CSV to S3 (filename includes date)
4) Create Snowflake table based on CSV header (all VARCHAR)
5) COPY INTO Snowflake table from S3

S3 output:
  s3://neighborwise-ai-s3-bucket/proximity/healthcare/healthcare_facilities_YYYY-MM-DD.csv

Snowflake:
  <SNOWFLAKE_DATABASE>.<SNOWFLAKE_SCHEMA>.STG_MA_HEALTHCARE_FACILITIES
"""

from __future__ import annotations

import os
import re
import io
import csv
import asyncio
import datetime as dt
from typing import Optional, List, Dict, Tuple


import boto3
from openpyxl import load_workbook
from dotenv import load_dotenv
from playwright.async_api import async_playwright

import snowflake.connector

from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.models import Variable
from airflow.utils.dates import days_ago


# -------------------------
# Load .env (optional)
# -------------------------
load_dotenv()


# =========================
# CONFIG
# =========================
PAGE_URL = "https://www.mass.gov/doc/list-of-health-care-facilities-licensed-or-certified-by-the-division"

AWS_REGION = os.getenv("AWS_DEFAULT_REGION", "us-east-2")
S3_BUCKET = Variable.get("s3_bucket", default_var="neighborwise-ai-s3-bucket")
S3_PREFIX = Variable.get("s3_prefix_healthcare", default_var="proximity/healthcare/")  # no dt folder

ROWS_TO_DROP = 8
SHEET_INDEX = 0

# Playwright
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/122.0.0.0 Safari/537.36"
)
HEADLESS = True          # in Docker this should be True
USE_SYSTEM_CHROME = False  # in Docker, use bundled chromium

# Snowflake target
SNOWFLAKE_DB = os.environ.get("SNOWFLAKE_DATABASE", "NEIGHBOURWISE_DOMAINS")
SNOWFLAKE_SCHEMA = os.environ.get("SNOWFLAKE_SCHEMA", "STAGE")
SNOWFLAKE_TABLE = os.environ.get("SNOWFLAKE_TABLE_HEALTHCARE", "STG_MA_HEALTHCARE_FACILITIES")
FULL_TABLE = f"{SNOWFLAKE_DB}.{SNOWFLAKE_SCHEMA}.{SNOWFLAKE_TABLE}"

# Local output dir inside container (safe writable path)
OUT_DIR = "/opt/airflow/data/_out_healthcare"


default_args = {
    "owner": "airflow",
    "depends_on_past": False,
    "retries": 1,
}


# =========================
# Helpers (Snowflake)
# =========================
def _sf_connect():
    """
    Uses env vars (as you already configured in docker-compose):
      SNOWFLAKE_ACCOUNT, SNOWFLAKE_USER, SNOWFLAKE_PASSWORD,
      SNOWFLAKE_WAREHOUSE, SNOWFLAKE_DATABASE, SNOWFLAKE_SCHEMA, SNOWFLAKE_ROLE (optional)
    """
    return snowflake.connector.connect(
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        user=os.environ["SNOWFLAKE_USER"],
        password=os.environ["SNOWFLAKE_PASSWORD"],
        warehouse=os.environ["SNOWFLAKE_WAREHOUSE"],
        database=SNOWFLAKE_DB,
        schema=SNOWFLAKE_SCHEMA,
        role=os.environ.get("SNOWFLAKE_ROLE"),
        insecure_mode=True,  # ok for dev; remove in prod
    )


def _sanitize_col(col: str) -> str:
    """Return a safe, uppercase Snowflake column identifier."""
    col = (col or "").strip()
    col = re.sub(r"[^0-9A-Za-z_]", "_", col)
    col = re.sub(r"__+", "_", col)
    col = col.strip("_").upper()
    if not col:
        col = "COL"
    if re.match(r"^\d", col):
        col = f"C_{col}"
    return col[:255]


def _dedupe_cols(cols: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out


# =========================
# Helpers (AWS/S3)
# =========================
def upload_bytes_to_s3(data: bytes, bucket: str, key: str, content_type: str) -> None:
    s3 = boto3.client("s3", region_name=AWS_REGION)  # uses env vars automatically
    s3.put_object(Bucket=bucket, Key=key, Body=data, ContentType=content_type)


# =========================
# Task 1: Download XLSX
# =========================
async def _download_massgov_xlsx_async() -> tuple[bytes, str]:
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context(
            user_agent=USER_AGENT,
            locale="en-US",
            accept_downloads=True,
            viewport={"width": 1280, "height": 800},
        )
        page = await context.new_page()

        try:
            resp = await page.goto(PAGE_URL, wait_until="domcontentloaded", timeout=90000)
            if resp and resp.status == 403:
                raise RuntimeError("Mass.gov page blocked (403) from this network/IP.")

            # Cookie banner best-effort
            for label in ["Accept", "Agree", "Accept all"]:
                try:
                    btn = page.get_by_role("button", name=re.compile(label, re.I))
                    await btn.click(timeout=2000)
                    break
                except Exception:
                    pass

            link = page.locator("a.ma__download-link__file-link").first
            await link.wait_for(state="visible", timeout=60000)

            href = await link.get_attribute("href")
            if not href:
                raise RuntimeError("Download link href is empty")

            download_url = ("https://www.mass.gov" + href) if href.startswith("/") else href

            async with page.expect_download(timeout=90000) as dl_info:
                try:
                    await page.goto(download_url, timeout=90000)
                except Exception as e:
                    # expected when navigation becomes a download
                    if "Download is starting" not in str(e):
                        raise

            download = await dl_info.value
            suggested = download.suggested_filename or "healthcare_facilities.xlsx"
            tmp_path = await download.path()
            if not tmp_path:
                raise RuntimeError("Download path empty (download not completed).")

            with open(tmp_path, "rb") as f:
                xlsx_bytes = f.read()

            return xlsx_bytes, suggested
        finally:
            await browser.close()


def task_download_xlsx(**context):
    os.makedirs(OUT_DIR, exist_ok=True)

    xlsx_bytes, suggested = asyncio.run(_download_massgov_xlsx_async())

    raw_path = os.path.join(OUT_DIR, "healthcare_facilities_raw.xlsx")
    with open(raw_path, "wb") as f:
        f.write(xlsx_bytes)

    context["ti"].xcom_push(key="raw_xlsx_path", value=raw_path)
    context["ti"].xcom_push(key="suggested_filename", value=suggested)

    return {"raw_xlsx_path": raw_path, "suggested_filename": suggested}


# =========================
# Task 2: XLSX -> Clean CSV
# =========================
def task_convert_xlsx_to_clean_csv(**context):
    ti = context["ti"]
    raw_path = ti.xcom_pull(task_ids="download_xlsx", key="raw_xlsx_path")
    if not raw_path or not os.path.exists(raw_path):
        raise FileNotFoundError(f"Missing raw XLSX at {raw_path}")

    # Read XLSX
    with open(raw_path, "rb") as f:
        xlsx_bytes = f.read()

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.worksheets[SHEET_INDEX]

    rows = list(ws.iter_rows(values_only=True))
    rows = rows[ROWS_TO_DROP:]  # drop top rows

    cleaned_rows = []
    for r in rows:
        if r is None:
            continue
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        cleaned_rows.append(r)

    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    for r in cleaned_rows:
        writer.writerow(["" if c is None else c for c in r])

    csv_text = out.getvalue()
    csv_bytes = csv_text.encode("utf-8")

    csv_path = os.path.join(OUT_DIR, "healthcare_facilities.csv")
    with open(csv_path, "wb") as f:
        f.write(csv_bytes)

    ti.xcom_push(key="clean_csv_path", value=csv_path)
    return {"clean_csv_path": csv_path, "rows": len(cleaned_rows)}


# =========================
# Task 3: Upload CSV to S3 (no dt folder; date in filename)
# =========================
def task_upload_csv_to_s3(**context):
    ti = context["ti"]
    csv_path = ti.xcom_pull(task_ids="convert_to_csv", key="clean_csv_path")
    if not csv_path or not os.path.exists(csv_path):
        raise FileNotFoundError(f"Missing cleaned CSV at {csv_path}")

    run_date = dt.date.today().isoformat()
    filename = f"healthcare_facilities_{run_date}.csv"
    s3_key = f"{S3_PREFIX}{filename}"

    with open(csv_path, "rb") as f:
        csv_bytes = f.read()

    upload_bytes_to_s3(csv_bytes, S3_BUCKET, s3_key, content_type="text/csv")

    ti.xcom_push(key="s3_bucket", value=S3_BUCKET)
    ti.xcom_push(key="s3_key", value=s3_key)
    ti.xcom_push(key="run_date", value=run_date)

    return {"s3_bucket": S3_BUCKET, "s3_key": s3_key}


# =========================
# Task 4: Create Snowflake table from CSV header
# =========================

from typing import List
import boto3, csv, io, re

def _read_csv_header_from_s3(bucket: str, key: str) -> List[str]:
    """
    Mass.gov CSV has a 2-line header (row 1 partial, row 2 completes words).
    This merges first two non-empty lines into one header row and returns
    sanitized + deduped Snowflake column names.
    """
    s3 = boto3.client("s3", region_name=AWS_REGION)
    obj = s3.get_object(Bucket=bucket, Key=key)

    # read enough bytes to safely include first 2 lines
    sample = obj["Body"].read(65536).decode("utf-8", errors="replace")
    lines = [ln for ln in sample.splitlines() if ln.strip()]

    if len(lines) < 2:
        raise ValueError("Could not read 2 header lines from CSV in S3")

    row1 = next(csv.reader(io.StringIO(lines[0])))
    row2 = next(csv.reader(io.StringIO(lines[1])))

    # pad to same length
    max_len = max(len(row1), len(row2))
    row1 += [""] * (max_len - len(row1))
    row2 += [""] * (max_len - len(row2))

    merged_raw = []
    for c1, c2 in zip(row1, row2):
        c1 = (c1 or "").strip()
        c2 = (c2 or "").strip()

        # If line1 is blank or "Unnamed", use line2
        if not c1 or c1.lower().startswith("unnamed"):
            merged = c2
        else:
            # If line2 adds meaning, append it
            if c2 and c2.lower() not in c1.lower():
                merged = f"{c1} {c2}"
            else:
                merged = c1

        merged_raw.append(merged)

    cols = [_sanitize_col(c) for c in merged_raw]
    cols = _dedupe_cols(cols)
    return cols


def task_create_snowflake_table(**context):
    ti = context["ti"]
    bucket = ti.xcom_pull(task_ids="upload_csv_to_s3", key="s3_bucket")
    key = ti.xcom_pull(task_ids="upload_csv_to_s3", key="s3_key")
    if not bucket or not key:
        raise ValueError("Missing s3_bucket/s3_key from XCom")

    cols = _read_csv_header_from_s3(bucket, key)

    # Option B: keep everything VARCHAR to avoid COPY failures
    col_defs = [f'"{c}" VARCHAR' for c in cols]

    # metadata columns
    col_defs += [
        "LOAD_TIMESTAMP TIMESTAMP_NTZ DEFAULT CURRENT_TIMESTAMP()",
        "_LOAD_DATE DATE DEFAULT CURRENT_DATE()",
        "_FILE_NAME VARCHAR",
    ]

    ddl = f"""
CREATE TABLE IF NOT EXISTS {FULL_TABLE} (
  {", ".join(col_defs)}
);
""".strip()

    conn = _sf_connect()
    cur = conn.cursor()
    try:
        cur.execute(ddl)
        return {"status": "table_ready", "table": FULL_TABLE, "columns": len(cols)}
    finally:
        cur.close()
        conn.close()


# =========================
# Task 5: COPY INTO Snowflake from S3
# =========================
def task_copy_into_snowflake(**context):
    ti = context["ti"]
    bucket = ti.xcom_pull(task_ids="upload_csv_to_s3", key="s3_bucket")
    key = ti.xcom_pull(task_ids="upload_csv_to_s3", key="s3_key")
    if not bucket or not key:
        raise ValueError("Missing s3_bucket/s3_key from XCom")

    aws_key = os.environ.get("AWS_ACCESS_KEY_ID")
    aws_secret = os.environ.get("AWS_SECRET_ACCESS_KEY")
    if not aws_key or not aws_secret:
        raise ValueError("AWS credentials not found in env inside Airflow container")

    # Build stage URL from prefix
    key_parts = key.split("/")
    filename = key_parts[-1]
    prefix = "/".join(key_parts[:-1]) + "/" if len(key_parts) > 1 else ""
    stage_url = f"s3://{bucket}/{prefix}"

    stage_name = "TMP_HEALTHCARE_S3_STAGE"
    file_format_name = "TMP_HEALTHCARE_CSV_FMT"

    create_stage_sql = f"""
CREATE OR REPLACE TEMP STAGE {stage_name}
URL = '{stage_url}'
CREDENTIALS = (AWS_KEY_ID='{aws_key}' AWS_SECRET_KEY='{aws_secret}');
""".strip()

    create_file_format_sql = f"""
CREATE OR REPLACE TEMP FILE FORMAT {file_format_name}
TYPE = CSV
FIELD_DELIMITER = ','
SKIP_HEADER = 2
FIELD_OPTIONALLY_ENCLOSED_BY = '"'
TRIM_SPACE = TRUE
NULL_IF = ('NULL', 'null', '')
EMPTY_FIELD_AS_NULL = TRUE
SKIP_BLANK_LINES = TRUE
ERROR_ON_COLUMN_COUNT_MISMATCH = FALSE;
""".strip()

    copy_sql = f"""
COPY INTO {FULL_TABLE}
FROM @{stage_name}/{filename}
FILE_FORMAT = (FORMAT_NAME = {file_format_name})
ON_ERROR = 'CONTINUE';
""".strip()

    # Optional: tag rows with filename for this run (if your table has _FILE_NAME)
    # This updates only rows loaded "today" that still have NULL _FILE_NAME.
    update_filename_sql = f"""
UPDATE {FULL_TABLE}
SET _FILE_NAME = '{filename}'
WHERE _LOAD_DATE = CURRENT_DATE()
  AND (_FILE_NAME IS NULL OR _FILE_NAME = '');
""".strip()

    # More meaningful validation than total count
    count_today_sql = f"""
SELECT COUNT(*) 
FROM {FULL_TABLE}
WHERE _LOAD_DATE = CURRENT_DATE();
""".strip()

    conn = _sf_connect()
    cur = conn.cursor()
    try:
        cur.execute(create_stage_sql)
        cur.execute(create_file_format_sql)
        cur.execute(copy_sql)

        # best-effort populate _FILE_NAME (won't fail if column missing? it WILL fail if missing)
        # So we guard it with a try.
        try:
            cur.execute(update_filename_sql)
        except Exception as e:
            print("[WARN] Could not update _FILE_NAME (column may not exist).", str(e))

        cur.execute(count_today_sql)
        count_today = cur.fetchone()[0]

        return {
            "status": "loaded",
            "table": FULL_TABLE,
            "rows_loaded_today": count_today,
            "s3_key": key,
            "stage_url": stage_url,
            "filename": filename,
        }
    finally:
        cur.close()
        conn.close()


# =========================
# DAG
# =========================
with DAG(
    dag_id="healthcare_dataload_dag",
    description="Mass.gov healthcare XLSX -> cleaned CSV -> S3 -> Snowflake",
    default_args=default_args,
    start_date=days_ago(1),
    schedule="0 6 * * *",  # daily 6 AM
    catchup=False,
    max_active_runs=1,
    tags=["neighborwise", "healthcare", "massgov", "s3", "snowflake"],
) as dag:

    download_xlsx = PythonOperator(
        task_id="download_xlsx",
        python_callable=task_download_xlsx,
        provide_context=True,
    )

    convert_to_csv = PythonOperator(
        task_id="convert_to_csv",
        python_callable=task_convert_xlsx_to_clean_csv,
        provide_context=True,
    )

    upload_csv_to_s3 = PythonOperator(
        task_id="upload_csv_to_s3",
        python_callable=task_upload_csv_to_s3,
        provide_context=True,
    )

    create_snowflake_table = PythonOperator(
        task_id="create_snowflake_table",
        python_callable=task_create_snowflake_table,
        provide_context=True,
    )

    copy_into_snowflake = PythonOperator(
        task_id="copy_into_snowflake",
        python_callable=task_copy_into_snowflake,
        provide_context=True,
    )

    download_xlsx >> convert_to_csv >> upload_csv_to_s3 >> create_snowflake_table >> copy_into_snowflake